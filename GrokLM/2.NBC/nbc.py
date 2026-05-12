import os
from openpyxl import load_workbook

script_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(script_dir, "dataset.xlsx")


def xl_2_list(file):
    data = []  # Пустой список, для сохранения извлеченных из файла данные
    workbook = load_workbook(file)
    page = workbook.active
    for row in page.iter_rows(values_only=True):
        item = {
            "ID": row[0],
            "x1": row[1],
            "x2": row[2],
            "x3": row[3],
            "Class": row[4],
        }  # {'ID': 1, 'x1': 'Цифра', 'x2': 'АК-74', 'x3': 'Патруль', 'Class': 'Свой'}

        data.append(item)

    return data


def get_x_list(full_dataset):
    data = []
    for item in full_dataset:
        if isinstance(item["ID"], int):
            data.append(item)
    return data


def get_X(full_dataset):
    for item in full_dataset:
        if item["Class"] == "?":
            return item


def get_y(x_list):
    CLASSES = []
    for x in x_list:
        if x["Class"] not in CLASSES:
            CLASSES.append(x["Class"])
    return CLASSES


def sort_dataset(x_list, y_list):
    sorted_ds = []
    for y in y_list:
        y_n_list = []
        for x in x_list:
            if x["Class"] == y:
                y_n_list.append(x)
        sorted_ds.append(y_n_list)
    return sorted_ds


def calc_priori_P(sorted_ds, x_count):
    priori_P_list = {}
    for y_n in sorted_ds:
        for x in y_n:
            priori_P_list[x["Class"]] = round(len(y_n) / x_count, 2)

    return priori_P_list

# def calc_cond_P()


list_from_file = xl_2_list(file_path)
print("\n\tПОЛНЫЙ ДАТАСЕТ ИЗ ФАЙЛА:", len(list_from_file), "\n", list_from_file)

x_list = get_x_list(list_from_file)
print("\n\tИЗВЕСТНЫЕ ОБЪЕКТЫ:", len(x_list), "\n", x_list)

X = get_X(list_from_file)
print("\n\tИСКОМЫЙ ОБЪЕКТ:", len(X), "\n", X)

y_list = get_y(x_list)
print("\n\tКЛАССЫ ДАТАСЕТА:", len(y_list), "\n", y_list)


sorted_ds = sort_dataset(x_list, y_list)
print("\n\tОТСОРТИРОВАННЫЙ ДАТАСЕТ:", len(sorted_ds), "\n", sorted_ds)

priori_P = calc_priori_P(sorted_ds, len(x_list))
print("\n\tАПРИОРНЫЕ ВЕРОЯТНОСТИ:", len(priori_P), "\n", priori_P)


# TODO 1. ИЗВЛЕЧЬ ДАТАСЕТ ИЗ ФАЙЛА
# TODO 2. ИЗВЛЕЧЬ ИЗ ДАТАСЕТА ИЗВЕСТНЫЕ ОБЪЕКТЫ
# TODO 3. ИЗВЛЕЧЬ ИЗ ДАТАСЕТА НЕИЗВЕСТНЫЙ ОБЪЕКТ
# TODO 4. ПОЛУЧИТЬ КЛАССЫ ИЗ ИЗВЕСТНОГО ДАТАСЕТА
# TODO 5. ОТСОРТИРОВАТЬ ИЗВЕСТНЫЙ ДАТАСЕТ ПО КЛАССАМ
# TODO 6. РАССЧИТАТЬ АППРИОРНЫЕ ВЕРОЯТНОСТИ
# TODO 7. РАССЧИТАТЬ УСЛОВНЫЕ ВЕРОЯТНОСТИ
