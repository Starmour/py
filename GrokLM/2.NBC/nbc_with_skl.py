# TODO: Импортировать NumPy и Sklearn
# TODO: Переписать методы из nbc.py чтобы они возвращали не словари {}, а масиивы []
import os
from openpyxl import load_workbook
import numpy as np
from sklearn.naive_bayes import GaussianNB
from sklearn.preprocessing import OrdinalEncoder

script_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(script_dir, "dataset.xlsx")


def xl_2_list(file):
    data = []  # Пустой список, для сохранения извлеченных из файла данные
    workbook = load_workbook(file)
    page = workbook.active
    for row in page.iter_rows(values_only=True):
        items = [row[0], row[1], row[2], row[3], row[4]]
        # {'ID': 1, 'x1': 'Цифра', 'x2': 'АК-74', 'x3': 'Патруль', 'Class': 'Свой'}

        data.append(items)

    return data


def get_params(in_array):
    params = []
    for object in in_array:
        items = []
        if isinstance(object[0], int):
            items.append(object[1])
            items.append(object[2])
            items.append(object[3])
            params.append(items)
    return params


def get_classes(in_array):
    classes = []
    for item in in_array:
        if isinstance(item[0], int):
            classes.append(item[-1])
    return classes


def get_X(in_array):
    obj = []
    for item in in_array:
        if item[4] == "?":
            obj.append(item[1])
            obj.append(item[2])
            obj.append(item[3])
    return obj


def coding_words(in_array):
    encoder = OrdinalEncoder()
    X = encoder.fit_transform(in_array)
    return X


dataset = xl_2_list(file_path)  # получаем список из файла
# print("XL_2_FILE:\n\t", dataset)

params = get_params(dataset)  # из списка получаем объекты с параметрами
# print("\n\tPARAMS:\n", X)
# print("\n\tPARAMS:\n", params)

y = np.array(get_classes(dataset))  # из списка получаем искомый объект
# print("\n\tCLASSES:\n", classes)

obj_x = get_X(dataset)  # из списка получаем искомый объект
# print("OBJ X:\n\t", obj_x)

params.append(obj_x)
x = coding_words(params)
print("\n\tENCODED PARAMETERS:\n", x)
x = np.delete(x, -1, axis=0)
print("\n\tENCODED PARAMETERS FROM DATASET:\n", x)
X = [x[-1]]
print("\n\tENCODED OBJECT_X PARAMETERS:\n", X)


model = GaussianNB()  # создаем обект класса GaussianNB
model.fit(x, y)


print("Априорные P(y):", model.class_prior_)
print("Средние (μ):", model.theta_)

print("Класс:", model.predict(X))
print("Вероятности:", model.predict_proba(X))